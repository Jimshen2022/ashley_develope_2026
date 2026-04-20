from openpyxl import load_workbook

wb = load_workbook(r'd:\python_file\AshtonOH3.xlsx')
ws1 = wb.worksheets
ws = ws1[0]
# ws.move_range("d4:f10",rows=-1,cols=2)
#
for row in ws['B2:G3']:
    for cell in row:
        print(cell.value)

# ws.merge_cells('c3:e8')
ws.unmerge_cells('c3:e8')
wb.save(r'd:\python_file\AshtonOH3.xlsx')





