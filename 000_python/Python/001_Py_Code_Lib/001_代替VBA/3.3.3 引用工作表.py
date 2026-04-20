import openpyxl as pyxl

wb = pyxl.Workbook()
ws = wb.create_sheet('Jim', 0)
sheets = wb.worksheets
print(len(sheets))

for sheet in sheets:
    print(sheet.title)

ws = sheets[0]
print(ws.title)

ws2 = wb['Jim']
print(ws2.title,'{}'.format(ws2.title))

ws3 = wb['Sheet']
# ws4 = wb.get_sheet_by_name('Jim')   ???
print(ws3.title)

names = wb.sheetnames
ws4 = wb[names[0]]
print(ws4.title)


















