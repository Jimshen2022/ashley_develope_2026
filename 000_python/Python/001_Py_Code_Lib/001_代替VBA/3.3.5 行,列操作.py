# 1 新增行
import openpyxl as pyxl

wb = pyxl.Workbook()
ws = wb.create_sheet('Myfile', 0)
# 获取所有表名
sheet_names = wb.sheetnames

ws = wb.create_sheet("Jim")
ws.append([10, 8, 21])
ws.append(['唐云', 39, 65])
ws.append({'a': '李广', 'b': 90, 'c': 87})
ws.append({1: '孙琦', 2: 83, 3: 79})

# 使用循环连续添加行数据
for row in range(1, 10):
    ws.append(range(10, 20))

# 获取行和列
row10 = ws[10]  # 第10行“有内容”的单元格
colC = ws['c']  # C列“有内容”的单元格

# 多行与多列的引用语法如下：
rows1 = ws[5:10]
cols1 = ws['c:d']

# 遍历行与列
# 遍历第1行的每个单元格
for cell in ws['1']:
    print(cell.value)

# 遍历第1列
for cell in ws['a']:
    print(cell.value)

# 遍历1-3行
for row in ws[1:3]:
    for cell in row:
        print(cell.value)

# 遍历1-3列
for column in ws['A:C']:
    for cell in column:
        print(cell.value)


# 遍历区域数据

for row in ws['a1:c3']:      # 遍历区域内的行
    for cell in row:        # 遍历每行的单元格
        print(cell.value)


# 将指定区域内的数据保存到列表DATA中, 并输出数据
data = []
for row in ws['a1:c3']:
    rv = []
    for cell in row:
        rv.append(cell.value)
        data.append(rv)
print(data)

# 获取包含工作表中所有数据的最小区域
wb2 = pyxl.load_workbook(r'd:\python_file\addrow.xlsx')
ws2 = wb2.active
print(ws2.min_row,ws2.max_row,ws2.min_column,ws2.max_column)

# 使用工作表对象的iter_rows, iter_cols方法，可以遍历指定区域内的行与列

for row in ws.iter_rows(min_row=2,max_col=10,max_row=3000):
    line = [cell.value for cell in row]
    print(line)

for col in ws.iter_cols(min_row=2,max_col=10,max_row=3000):
    line1 = [cell.value for cell in col]
    print(line1)

print('------------------------------------------------------------------------------------------------------')
# 遍历所有行或列
for row in ws.rows:
    line2 = [cell.value for cell in row]
    print(line2)
print('------------------------------------------------------------------------------------------------------')

for col in ws.columns:
    line3 = [cell.value for cell in col]
    print(line3)
print('*******************************************************************************************************')

# 使用工作表对象的values属性返回各行的数据
for row in ws.values:
    print(row)

# 以列表的形式输出每行的数据
for row in ws.values:
    print(list(row))


# 6. 插入和删除行
ws.insert_rows(5)    # 在第5行上面插入一个空行
ws.insert_rows(5,3)    # 在第5行上面插入3个空行

ws.insert_cols(4)     # 第4列左侧插入1列
ws.insert_cols(4,3)    # 第4列左侧插入3列

ws.delete_rows(5,3)    # 从第5行开始, 连续删除3行
ws.delete_cols(4,3)    # 从第4列开始, 连续删除3行


# 7 改变行高与列宽
ws.row_dimensions[2].height = 20
ws.column_dimensions['C'].width = 35










wb.save(r'd:\python_file\addrow.xlsx')







