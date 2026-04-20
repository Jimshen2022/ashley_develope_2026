import pyodbc as po
import pandas as pd
import os
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor, as_completed

# ===== 配置数据库连接 =====
CONN_STR = 'DSN=AFIPROD;UID=JIMSHEN;PWD=MJ2083'

# 表与 schema 映射
TABLES = {
    'TSININA1': 'AFILELIB',
    'TSITXN': 'AFILELIB',
    'MBBZRES1': 'AMFLIBA',  # 特殊 schema
    'TSITIN': 'AFILELIB',
    'TSINXN': 'AFILELIB',
    'ITMEXT': 'AFILELIB',
    'ACUSMASJ': 'AFILELIB'
}

MAX_ROWS = 1000000
MAX_WORKERS = 4  # 并发线程数

# ===== 样式 =====
BOLD = Font(bold=True)
GRAY_FILL = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
CENTER = Alignment(horizontal='center', vertical='center')
DESC_FONT = Font(color="666666")

# ===== 工具函数 =====
def fetch_data(query):
    cnxn = po.connect(CONN_STR, autocommit=True)
    df = pd.read_sql(query, cnxn)
    cnxn.close()
    return df

def get_column_desc(schema, table):
    query = f"""
        SELECT COLUMN_NAME, VARCHAR(COLUMN_TEXT, 100) AS COLUMN_DESC
        FROM QSYS2.SYSCOLUMNS
        WHERE SYSTEM_TABLE_SCHEMA = '{schema}' AND TABLE_NAME = '{table}'
        ORDER BY ORDINAL_POSITION
    """
    df = fetch_data(query)
    return dict(zip(df['COLUMN_NAME'], df['COLUMN_DESC']))

def process_table(table, schema):
    full_name = f"{schema}.{table}"
    try:
        print(f"📥 正在抓取：{full_name}")
        df = fetch_data(f"SELECT * FROM {full_name} FETCH FIRST {MAX_ROWS} ROWS ONLY")
        desc_map = get_column_desc(schema, table)
        return (full_name, df, desc_map)
    except Exception as e:
        print(f"❌ 抓取失败：{full_name} — {e}")
        return None

def write_sheet(wb, sheet_name, df, column_desc_map):
    ws = wb.create_sheet(title=sheet_name)
    columns = df.columns.tolist()
    descs = [column_desc_map.get(col, '') for col in columns]

    # 第1行：列名
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = BOLD
        cell.alignment = CENTER
        cell.fill = GRAY_FILL

    # 第2行：字段说明
    for col_idx, desc in enumerate(descs, 1):
        cell = ws.cell(row=2, column=col_idx, value=desc)
        cell.font = DESC_FONT
        cell.alignment = CENTER

    # 冻结前2行
    ws.freeze_panes = "A3"

    # 第3行起：数据
    for row_idx, row in enumerate(df.itertuples(index=False, name=None), start=3):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 自动列宽（使用 get_column_letter 修复）
    for col_idx, col_name in enumerate(columns, start=1):
        max_len = max(
            len(str(col_name)),
            len(str(descs[col_idx - 1])) if descs[col_idx - 1] else 0,
            *(len(str(r[col_idx - 1])) for r in df.itertuples(index=False, name=None))
        )
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

# ===== 主程序 =====
def main():
    start = time.time()
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认 Sheet

    print("🚀 多线程开始抓取数据...\n")
    results = []

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {
            executor.submit(process_table, table, schema): (schema, table)
            for table, schema in TABLES.items()
        }

        for future in as_completed(futures):
            result = future.result()
            if result:
                results.append(result)

    for sheet_name, df, desc_map in results:
        write_sheet(wb, sheet_name, df, desc_map)

    # 输出路径
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    file_name = f"AFI Shipping file_{timestamp}.xlsx"
    downloads_path = os.path.join(os.path.expanduser('~'), 'Downloads')
    file_path = os.path.join(downloads_path, file_name)

    wb.save(file_path)
    print(f"\n✅ 文件已保存：{file_path}")
    print(f"⏱ 总耗时：{time.time() - start:.2f} 秒")

if __name__ == '__main__':
    main()
