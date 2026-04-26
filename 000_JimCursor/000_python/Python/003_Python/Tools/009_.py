import fitz  # PyMuPDF
import os
import re
import openpyxl

# 定义要提取的字段名称
fields = [
    "BOOKING NUMBER:",
    "TOTAL BOOKING CONTAINER QTY SIZE/TYPE:",
    "INTENDED VESSEL/VOYAGE:",
    "ETD:",
    "FINAL DESTINATION:",
    "ETA:",
    "INTENDED VGM CUT-OFF:",
    "INTENDED FCL CY CUT-OFF:",
    "INTENDED ESI CUT-OFF:"
]

def extract_field_data_from_pdf(pdf_path):
    doc = fitz.open(pdf_path)
    field_data = {}
    current_field = None
    current_value = []
    collecting = False  # 标记是否在收集字段值

    # 定义正则表达式来匹配字段
    field_patterns = {field: re.compile(re.escape(field)) for field in fields}
    
    for page_num in range(doc.page_count):
        page = doc[page_num]
        lines = page.get_text("text").split("\n")
        
        for line in lines:
            line = line.strip()
            
            # 如果正在收集字段的值
            if collecting:
                # 检查当前行是否是下一个字段
                if any(pattern.match(line) for pattern in field_patterns.values()):
                    # 遇到下一个字段，停止收集当前字段的值
                    field_data[current_field] = " ".join(current_value).strip()
                    current_field = None
                    current_value = []
                    collecting = False  # 停止收集
                else:
                    # 否则继续收集当前字段的多行值
                    current_value.append(line)
            
            # 如果不在收集状态，检查这一行是否是一个新的字段
            if not collecting:
                for field, pattern in field_patterns.items():
                    match = pattern.match(line)
                    if match:
                        # 找到新的字段，开始记录它的值
                        current_field = field
                        value = line[len(field):].strip()  # 提取字段后的内容
                        current_value.append(value)
                        collecting = True  # 开始收集字段值
                        break
    
    # 如果最后还有未结束的字段，保存它的值
    if current_field is not None:
        field_data[current_field] = " ".join(current_value).strip()
    
    doc.close()
    return field_data

def write_to_excel(data, pdf_filename, output_excel_path):
    # 如果包含"中远"，则放到 "中远" sheet 中，否则创建以文件名命名的 sheet
    if "中远" in data.get("INTENDED VESSEL/VOYAGE:", ""):
        sheet_name = "中远"
    else:
        sheet_name = pdf_filename.replace('.pdf', '')[:31]  # Excel sheet name 长度不能超过31
    
    # 如果 Excel 文件不存在则创建
    if not os.path.exists(output_excel_path):
        workbook = openpyxl.Workbook()
    else:
        workbook = openpyxl.load_workbook(output_excel_path)

    # 如果 sheet 已存在则使用，否则创建新 sheet
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)
        # 添加表头
        sheet.append(["PDF Filename"] + fields)

    # 创建新行并添加文件名及字段值
    row = [pdf_filename]
    for field in fields:
        row.append(data.get(field, ""))  # 如果字段不存在，留空

    # 添加这一行数据
    sheet.append(row)
    
    # 保存 Excel 文件
    workbook.save(output_excel_path)
    print(f"Data written to sheet: {sheet_name} in {output_excel_path}")

# 主函数
def process_pdfs_in_folder(folder_path, output_excel_path):
    for filename in os.listdir(folder_path):
        if filename.endswith(".pdf"):
            print(f"Processing file: {filename}")
            pdf_path = os.path.join(folder_path, filename)
            field_data = extract_field_data_from_pdf(pdf_path)
            write_to_excel(field_data, filename, output_excel_path)

# 运行代码
folder_path = '/Python2038/016_Python/Tools/pdf'  # 替换为你的 PDF 文件夹路径
output_excel_path = '/Python2038/016_Python/Tools/output.xlsx'  # 输出的 Excel 文件
process_pdfs_in_folder(folder_path, output_excel_path)
