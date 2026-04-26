import openpyxl
from openpyxl.styles import numbers, Font, Alignment
from openpyxl.styles import PatternFill, Border, Side, Protection
from openpyxl.styles import Font
from openpyxl.styles.colors import Color
from openpyxl.styles import GradientFill


wb = openpyxl.load_workbook(r'd:\python_file\AshtonOH3.xlsx')
ws1 = wb.worksheets
ws = ws1[0]

# 1.设置字体
font = Font(name='Arial', size=14, bold=True, italic=True,
            underline='single', strike=False, color='FF0000')

# 将C3 value设为粗体
# font = Font(bold=True)
# cl = ws['c3']
# cl.font = font

# 遍历单元格，并设置字体
# for row in ws['a1:g30']:
#     for cell in row:
#         cell.font = font

# row5 = ws.dimensions
# row0 = ws.row_dimensions[0]

# 设置某个单元格格式
ws.cell(3,3).font = font
ws['a1'].font = font

# 遍历range 并设置字体
# for row in ws['a1:g2']:
#     for cell in row:
#         cell.font = font


#  遍历有资料的range 并设置颜色
# font = Font(color='00FF0000')
# for rows in ws[ws.dimensions]:
#     for cell in rows:
#         cell.font = font

# 设置行高,列宽
ws.row_dimensions[1].height = 30
# ws.column_dimensions['A'].width = 30

# 设置颜色
# 16进制RGB
font2 = Font(color='00FF0000')
font3 = Font(color='FF0000')

c = Color(rgb='00FF00')   # RGB颜色
font4 = Font(color=c)

c2 = Color(index=32)  # 索引着色
font5 = Font(color=c2)

c3 = Color(theme=6, tint=0.5)    # 主题颜色
font6 = Font(color=c3)

# 遍历range 并设置字体
# for row in ws['a1:g3']:
#     for cell in row:
#         cell.font = font6

# 3.样式--设置背景填充
# (1) 渐变色填充
# ws['B6'].fill = GradientFill(type='linear',degree=0,left=0,right=0,
#                                   top=0,bottom=0,stop=['FF0000','0000FF'])
#
# ws['E6'].fill = GradientFill(type='linear',degree=45,left=0,right=0,
#                                   top=0,bottom=0,stop=['FF0000','0000FF'])
#
# ws['G6'].fill = GradientFill(type='path',left=0.2,right=0.8,
#                                   top=0.3,bottom=0.7,stop=['FF0000','0000FF'])
#
# ws.row_dimensions[14].fill = GradientFill(type='linear',degree=45,left=0,right=0,
#                                   top=0,bottom=0,stop=['FF0000','00FF00'])
#
#
# ws.row_dimensions[20].fill = GradientFill(type='linear',degree=45,left=0,right=0,
#                                   top=0,bottom=0,stop=['FF0000','00FF00'])


# 有资料区域加背景填充
# for rows in ws[ws.dimensions]:
#     for cell in rows:
#         cell.fill = GradientFill(type='linear',degree=45,left=0,right=0,
#                                   top=0,bottom=0,stop=['FF0000','00FF00'])


# （2）图案填充
# from openpyxl.styles import PatternFill
# ws['b12'].fill = PatternFill(fill_type=None, start_color='FFFF00',end_color='000000')
# ws['c12'].fill = PatternFill(fill_type='solid', start_color='00FF00')
# ws['e12'].fill = PatternFill(fill_type='lightGrid', start_color='FFFF00',end_color='000000')


# 指定第2列的背景色
#
# fill = PatternFill(fill_type='lightTrellis', fgColor=Color(rgb='00FF00'), bgColor=Color(rgb='0000FF'))
# ws.column_dimensions['B'].fill = fill

# 指定第6行的背景色

# fill = PatternFill(fill_type='lightTrellis', fgColor=Color(rgb='00FF00'), bgColor=Color(rgb='0000FF'))
# ws.row_dimensions[6].fill = fill

# 4.设置边框
from openpyxl.styles import Border, Side
ws.cell(row=4,column=4) .border =Border(left=Side(border_style='thin',color='FF0000'),
                                right=Side(border_style='thin',color='00FF00'),
                                top=Side(border_style='double',color='FF0000'),
                                bottom=Side(border_style='double',color='FF0000'))

# 5.设置数字格式
# from openpyxl.styles import numbers
# ws['B5'].number_format = numbers.FORMAT_GENERAL
# ws['B6'].number_format = "yy-mm-dd"
# ws['B6'].number_format = "d-mmm-yy"
# ws['B6'].number_format = "0.00E+00"

# 6. 设置对齐方式
from openpyxl.styles import Alignment

align1 = Alignment(horizontal='center',vertical='top')
align2 = Alignment(horizontal='right',vertical='bottom',text_rotation=30,wrap_text=True,
                   shrink_to_fit=True, indent=0)
align3 = Alignment(horizontal='center',vertical='center',wrap_text=True, indent=3)

ws['c2'].alignment = align1
ws['c4'].alignment = align2
ws['c6'].alignment = align3

# 7.设置保护
from openpyxl.styles import Protection
ws['c3'].protection = Protection(locked=True,hidden=False)








wb.save(r'd:\python_file\AshtonOH3.xlsx')
