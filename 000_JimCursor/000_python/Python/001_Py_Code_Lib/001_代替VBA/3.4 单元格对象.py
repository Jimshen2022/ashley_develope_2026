# 3.4.1 单元格的引用与赋值
from openpyxl import load_workbook

wb = load_workbook(r'd:\python_file\AshtonOH.xlsx')
ws = wb.worksheets[0]
cl0 = ws['a1']
cl1 = ws['a1'].internal_value
cl2 = ws.cell(row=2,column=2,value=10)
print(cl0)
print(cl1)
print(cl2)

# cell对象的主要属性和方法
c1 = ws['c3']
print(c1.row)    # 单元格的行号
print(c1.column)    # 单元格的行号
print(c1.value)    # 单元格的值
print(c1.coordinate)    # 单元格的坐标
print(c1.data_type)    # 单元格中数据类型
# c1.hyperlink.ref="https:\\www.baidu.com"    # 单元格的链接 ??
# print(c1.hyperlink)
h = c1.offset(row=1,column=2)
print(h.value)









