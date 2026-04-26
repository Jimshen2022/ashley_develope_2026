import openpyxl
from openpyxl import load_workbook

# 文件路径
file_path = r"C:\Users\jishen\Downloads\RP Unloaded from trip 12608.xlsx"

# 加载工作簿
wb = load_workbook(file_path)

# 删除旧的 Merged 表（如果已存在）
if "Merged" in wb.sheetnames:
    del wb["Merged"]

# 创建合并表
merged_sheet = wb.create_sheet("Merged")

# 获取原始标题 + 加上“Sheet Name”列标题
for sheet_name in wb.sheetnames:
    if sheet_name != "Merged":
        header_row = [cell.value for cell in wb[sheet_name][1][:4]]  # A~D列标题
        header_row.append("Sheet Name")  # 添加新列标题
        merged_sheet.append(header_row)
        break

# 遍历工作表
for sheet_name in wb.sheetnames:
    if sheet_name == "Merged":
        continue

    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
        if row[0] is None:  # A列为空
            break
        row_with_sheet = list(row) + [sheet_name]  # 添加 Sheet 名称
        merged_sheet.append(row_with_sheet)

# 保存结果
wb.save(r"C:\Users\jishen\Downloads\RP Merged.xlsx")

print("合并完成，每行已附带 Sheet 名称，文件保存为 RP Merged.xlsx")
