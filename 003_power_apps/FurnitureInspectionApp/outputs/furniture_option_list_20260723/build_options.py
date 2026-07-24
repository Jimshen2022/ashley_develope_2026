from collections import defaultdict
from pathlib import Path
import re
import unicodedata

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

INPUT = Path(r"C:\Users\jishen\Downloads\list.xlsx")
OUTPUT = Path(r"D:\GitHub\ashley_develope_2026\003_power_apps\FurnitureInspectionApp\outputs\furniture_option_list_20260723\FurnitureInspectionOptionList.xlsx")

CATEGORY_MAP = {
    "Damaged Description": "DamagedDescription",
    "Reason": "Reason",
    "Damaged by": "DamagedBy",
    "Whse deal with status": "WhseDealWithStatus",
}

MANUAL_MAP = {
    "DamagedDescription": {
        "carton damage": "Carton Damage",
        "damaged carton": "Carton Damage",
        "water-damaged carton": "Water-damaged carton",
        "water damaged carton": "Water-damaged carton",
        "water on carton": "Water-damaged carton",
        "dent on the headrest": "Dent on headrest",
        "dent on headrest": "Dent on headrest",
        "scraped wood": "Scraped Wood",
    },
    "Reason": {
        "damage found at storage location": "Damage found at storage location",
        "damage found at storage inspection": "Damage found at storage location",
        "damage found in pick": "Damage found during picking",
        "damage found during picking": "Damage found during picking",
        "crash during picking": "Damage found during picking",
        "handling in location": "Handling in location",
        "handling in location": "Handling in location",
        "vendor mistake": "Vendor mistake",
        "sent by vendor": "Vendor mistake",
        "vendor shipped": "Vendor shipped",
    },
    "DamagedBy": {
        "haven't found the cause yet": "Cause not yet found",
        "ashton havent found root cause yet": "Cause not yet found",
        "ashton haven't found root cause yet": "Cause not yet found",
        "vendor error": "Vendor error",
        "vendor": "Vendor error",
        "vo minh cuong": "Vo Minh Cuong",
        "phan văn danh": "PHAN VĂN DANH",
    },
    "WhseDealWithStatus": {
        "need get materials from vendor then we can repair": "Need materials from vendor then we can repair",
        "we cannot repair it need return to vendor for repair": "Cannot repair, return to vendor for repair",
        "we cannot repair it, need return to vendor for repair": "Cannot repair, return to vendor for repair",
    },
}


def clean_text(value):
    if pd.isna(value):
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    text = text.replace("’", "'").replace("`", "'")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def key_text(text):
    text = clean_text(text).lower()
    text = re.sub(r"[.]+$", "", text).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def title_case_safe(text):
    if not text:
        return text
    if text.isupper() and len(text) > 4 and not any(ch in text for ch in "ĂÂÊÔƠƯĐ"):
        return text.title()
    return text[0].upper() + text[1:]


def canonical(category, value):
    text = clean_text(value)
    key = key_text(text)
    mapped = MANUAL_MAP.get(category, {}).get(key)
    if mapped:
        return mapped
    return title_case_safe(text)


df = pd.read_excel(INPUT, dtype=str)
counts = defaultdict(lambda: {"count": 0, "originals": defaultdict(int)})

for source_col, category in CATEGORY_MAP.items():
    if source_col not in df.columns:
        continue
    for raw in df[source_col].tolist():
        text = clean_text(raw)
        if not text:
            continue
        name = canonical(category, text)
        counts[(category, name)]["count"] += 1
        counts[(category, name)]["originals"][text] += 1

sharepoint_rows = []
audit_rows = []
for category in CATEGORY_MAP.values():
    cat_items = [
        (name, data["count"], data["originals"])
        for (cat, name), data in counts.items()
        if cat == category
    ]
    cat_items.sort(key=lambda x: (-x[1], x[0].lower()))
    for idx, (name, count, originals) in enumerate(cat_items, start=1):
        sort_order = idx * 10
        sharepoint_rows.append([name, category, "Yes", sort_order])
        merged_from = " | ".join(
            f"{orig} ({orig_count})"
            for orig, orig_count in sorted(originals.items(), key=lambda kv: (-kv[1], kv[0].lower()))
        )
        audit_rows.append([name, category, "Yes", sort_order, count, merged_from])

wb = Workbook()
ws = wb.active
ws.title = "SharePointOptions"
ws.append(["Title", "Category", "Active", "SortOrder"])
for row in sharepoint_rows:
    ws.append(row)

ws2 = wb.create_sheet("AuditCounts")
ws2.append(["Title", "Category", "Active", "SortOrder", "Count", "MergedFrom"])
for row in audit_rows:
    ws2.append(row)

ws3 = wb.create_sheet("README")
readme = [
    ["Purpose", "Use SharePointOptions to create/maintain one central options list for Power Apps dropdowns."],
    ["Power Apps filter", "SortByColumns(Filter(FurnitureInspectionOptions, Category = \"Reason\" And Active = true), \"SortOrder\", SortOrder.Ascending)"],
    ["Categories", ", ".join(CATEGORY_MAP.values())],
    ["Merge rule", "Only obvious duplicates were merged: case, spacing, punctuation, clear spelling/capitalization variants, and very clear same-meaning phrases."],
    ["Active", "Use Yes/No column in SharePoint so users can deactivate options without deleting history."],
    ["SortOrder", "Sorted by frequency descending, then text ascending. Gaps of 10 leave room for manual inserts."],
]
for row in readme:
    ws3.append(row)

header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="D9E2F3")
border = Border(bottom=thin)

for sheet in [ws, ws2, ws3]:
    sheet.sheet_view.showGridLines = False
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    sheet.freeze_panes = "A2"

widths = {
    "SharePointOptions": [48, 24, 12, 12],
    "AuditCounts": [48, 24, 12, 12, 10, 90],
    "README": [22, 110],
}
for sheet in [ws, ws2, ws3]:
    for idx, width in enumerate(widths[sheet.title], start=1):
        sheet.column_dimensions[chr(64 + idx)].width = width

for sheet, table_name in [(ws, "SharePointOptionsTable"), (ws2, "AuditCountsTable")]:
    last_col = sheet.max_column
    last_row = sheet.max_row
    ref = f"A1:{chr(64 + last_col)}{last_row}"
    tab = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    sheet.add_table(tab)

ws.auto_filter.ref = f"A1:D{ws.max_row}"
ws2.auto_filter.ref = f"A1:F{ws2.max_row}"

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT)
print(OUTPUT)
print(f"SharePointOptions rows: {len(sharepoint_rows)}")
for category in CATEGORY_MAP.values():
    print(category, sum(1 for row in sharepoint_rows if row[1] == category))
