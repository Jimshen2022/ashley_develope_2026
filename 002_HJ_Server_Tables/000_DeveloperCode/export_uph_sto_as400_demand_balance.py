from __future__ import annotations

import argparse
import getpass
import os
import re
from datetime import date, datetime
from pathlib import Path

from openpyxl.styles import Font, PatternFill
import pandas as pd
import pyodbc


DEFAULT_SERVER = os.getenv("HJ_SERVER", "AshtonWHJSQLprod")
DEFAULT_DATABASE = os.getenv("HJ_DATABASE", "AAD")
DEFAULT_DRIVER = os.getenv("HJ_ODBC_DRIVER", "ODBC Driver 17 for SQL Server")
DEFAULT_WH_ID = os.getenv("HJ_WH_ID", "335")
DEFAULT_PICK_PUT_ID = os.getenv("HJ_PICK_PUT_ID", "UPH")
DEFAULT_ENCRYPT = os.getenv("HJ_ENCRYPT", "no")

DEFAULT_DB2_DSN = os.getenv("DB2_DSN", "AFIPROD")
DEFAULT_DB2_UID = os.getenv("DB2_UID", "JIMSHEN")

BUCKET_DAYS = [30, 60, 90, 180, 360]


def load_hj_sql_query(file_name: str) -> str:
    sql_path = Path(__file__).with_name(file_name)
    sql = sql_path.read_text(encoding="utf-8-sig")
    sql = sql.replace("DECLARE @wh_id VARCHAR(10) = '335';", "DECLARE @wh_id VARCHAR(10) = ?;")
    sql = sql.replace("DECLARE @pick_put_id VARCHAR(15) = 'UPH';", "DECLARE @pick_put_id VARCHAR(15) = ?;")
    return sql


def load_full_as400_demand_sql() -> str:
    sql_path = Path(__file__).with_name("db2_ashton_demand.sql")
    sql = sql_path.read_text(encoding="utf-8-sig")
    sql = "\n".join(line for line in sql.splitlines() if not line.lstrip().startswith("--")).strip().rstrip(";")
    sql = re.sub(
        r"ORDER\s+BY\s+t1\.BDTRP#\s*,\s*t1\.BDISEQ\s*,\s*t1\.BDITM#\s*\)\s*x1",
        ") x1",
        sql,
        flags=re.IGNORECASE,
    )
    sql = re.sub(
        r"ORDER\s+BY\s+a1\.MFIDT\s*,\s*x1\.BDTRP#\s*,\s*a1\.ITNBR\s*,\s*x1\.BDISEQ\s*$",
        "",
        sql,
        flags=re.IGNORECASE,
    ).strip()
    return sql


def build_light_as400_demand_sql() -> str:
    return """
SELECT
    a1.ORDNO,
    a1.ITMSQ,
    a1.CCUSNO,
    a1.CUSNM,
    a1.CSHPNO,
    x1.BDTRP#,
    x1.BDITQT AS TRIP_QTY,
    a1.HOUSE,
    a1.ITNBR,
    a1.OPEN_CO_QTY,
    a1.LoadDate
FROM (
    SELECT
        t1.HOUSE,
        t1.ORDNO,
        t1.ITMSQ,
        TRIM(t1.ITNBR) AS ITNBR,
        t1.CCUSNO,
        t3.CUSNM,
        t1.CSHPNO,
        CHAR(t1.MFIDT) AS LoadDate,
        t1.COQTY - t1.QTYSH AS OPEN_CO_QTY
    FROM AFILELIB.CODATAN t1, AFILELIB.EXTORD t2, AFILELIB.ACUSMASJ t3, AFILELIB.COMAST t4
    WHERE t2.XORDNO = t1.ORDNO
      AND t3.CUSNO = t1.CCUSNO
      AND t1.ORDNO = t4.ORDNO
      AND t1.HOUSE IN (?)
      AND t1.COQTY - t1.QTYSH <> 0
) a1
LEFT JOIN (
    SELECT
        t1.BDTRP#,
        t1.BDORD#,
        t1.BDISEQ,
        t1.BDITM#,
        t1.BDCUS#,
        t1.BDITQT
    FROM DISTLIB.BTTRIPD t1, DISTLIB.BTTRIPH t2
    WHERE t2.BHWHS# IN (?)
      AND t2.BHLDAT BETWEEN 0 AND 20261231
      AND t2.BHTRPS IN ('A','R','X')
      AND t1.BDTRP# = t2.BHTRP#
) x1 ON a1.ORDNO || a1.ITMSQ || a1.ITNBR || a1.CCUSNO = x1.BDORD# || x1.BDISEQ || x1.BDITM# || x1.BDCUS#
""".strip()

def build_hj_connection_string(args: argparse.Namespace) -> str:
    return (
        f"DRIVER={{{args.driver}}};"
        f"SERVER={args.server};"
        f"DATABASE={args.database};"
        "Trusted_Connection=yes;"
        f"Encrypt={args.encrypt};"
        "TrustServerCertificate=yes;"
    )


def build_db2_connection_string(args: argparse.Namespace) -> str:
    password = args.db2_password or os.getenv("DB2_PWD")
    if not password:
        password = getpass.getpass("AS400 DB2 password: ")
    return f"DSN={args.db2_dsn};UID={args.db2_uid};PWD={password}"


def default_output_path() -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Path.home() / "Downloads" / f"uph_sto_as400_demand_balance_{timestamp}.xlsx"


def normalize_column_name(df: pd.DataFrame, target: str) -> str:
    target_norm = target.lower().replace(" ", "").replace("_", "")
    for col in df.columns:
        col_norm = str(col).lower().replace(" ", "").replace("_", "")
        if col_norm == target_norm:
            return col
    raise KeyError(f"Column not found: {target}")


def clean_numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").fillna(0)


def parse_as400_yyyymmdd(series: pd.Series) -> pd.Series:
    text = series.astype("string").str.strip().str.replace(r"\.0$", "", regex=True)
    text = text.mask(text.isin(["", "0", "00000000", "<NA>"]))
    return pd.to_datetime(text, format="%Y%m%d", errors="coerce")


def yes_mask(series: pd.Series) -> pd.Series:
    return series.astype("string").str.upper().eq("Y")


def fetch_hj_stored_item_detail(args: argparse.Namespace) -> pd.DataFrame:
    query = load_hj_sql_query("uph_stored_item_detail.sql")
    with pyodbc.connect(build_hj_connection_string(args)) as conn:
        return pd.read_sql_query(query, conn, params=[args.wh_id, args.pick_put_id])


def fetch_as400_demand_detail(args: argparse.Namespace) -> pd.DataFrame:
    if args.full_db2_demand_detail:
        query = load_full_as400_demand_sql()
        params = None
    else:
        query = build_light_as400_demand_sql()
        params = [args.wh_id, args.wh_id]

    with pyodbc.connect(build_db2_connection_string(args), autocommit=True) as conn:
        df = pd.read_sql_query(query, conn, params=params)

    item_col = normalize_column_name(df, "ITNBR")
    qty_col = normalize_column_name(df, "OPEN_CO_QTY")
    load_date_col = normalize_column_name(df, "LOADDATE")

    df = df.copy()
    df["item_number"] = df[item_col].astype("string").str.strip()
    df["demand_qty"] = clean_numeric(df[qty_col])
    df["demand_date"] = parse_as400_yyyymmdd(df[load_date_col])


    df = df[df["demand_qty"] > 0].copy()
    df["demand_date_text"] = df["demand_date"].dt.strftime("%Y-%m-%d")
    sort_cols = [col for col in ["item_number", "demand_date", "ORDNO", "ITMSQ"] if col in df.columns]
    return df.sort_values(sort_cols, kind="stable") if sort_cols else df


def build_inventory_summary(stored_item_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    if stored_item_df.empty:
        empty_summary = pd.DataFrame(columns=["item_number", "sto_qty", "shipping_stage_qty"])
        empty_location = pd.DataFrame(columns=["item_number", "location_id", "location_qty"])
        return empty_summary, empty_location

    df = stored_item_df.copy()
    df["item_number"] = df["item_number"].astype("string").str.strip()
    df["location_id"] = df["location_id"].astype("string").str.strip()
    df["actual_qty"] = clean_numeric(df["actual_qty"])
    df["is_sto_qty_source_bool"] = yes_mask(df["is_sto_qty_source"])
    df["is_shipping_stage_bool"] = yes_mask(df["is_shipping_stage"])

    sto = df[df["is_sto_qty_source_bool"]].groupby("item_number", dropna=False)["actual_qty"].sum().rename("sto_qty")
    stage = df[df["is_shipping_stage_bool"]].groupby("item_number", dropna=False)["actual_qty"].sum().rename("shipping_stage_qty")
    summary = pd.concat([sto, stage], axis=1).fillna(0).reset_index()

    location = (
        df[df["is_sto_qty_source_bool"] & ~df["is_shipping_stage_bool"]]
        .groupby(["item_number", "location_id"], dropna=False)["actual_qty"]
        .sum()
        .rename("location_qty")
        .reset_index()
        .sort_values(["item_number", "location_id"], kind="stable")
    )
    return summary, location


def build_demand_summary(demand_df: pd.DataFrame, as_of_date: date, bucket_mode: str) -> pd.DataFrame:
    columns = ["item_number", "all_demand_qty"] + [f"demand_{days}_days" for days in BUCKET_DAYS] + ["demand_greater_than_360_day"]
    if demand_df.empty:
        return pd.DataFrame(columns=columns)

    df = demand_df.copy()
    df["item_number"] = df["item_number"].astype("string").str.strip()
    df["demand_qty"] = clean_numeric(df["demand_qty"])
    as_of_timestamp = pd.Timestamp(as_of_date)
    days_from_as_of = (df["demand_date"] - as_of_timestamp).dt.days

    summary = df.groupby("item_number", dropna=False)["demand_qty"].sum().rename("all_demand_qty").to_frame()
    if bucket_mode == "cumulative":
        bucket_masks = {
            "demand_30_days": df["demand_date"].isna() | (days_from_as_of <= 30),
            "demand_60_days": df["demand_date"].isna() | (days_from_as_of <= 60),
            "demand_90_days": df["demand_date"].isna() | (days_from_as_of <= 90),
            "demand_180_days": df["demand_date"].isna() | (days_from_as_of <= 180),
            "demand_360_days": df["demand_date"].isna() | (days_from_as_of <= 360),
            "demand_greater_than_360_day": days_from_as_of > 360,
        }
    else:
        bucket_masks = {
            "demand_30_days": df["demand_date"].isna() | (days_from_as_of <= 30),
            "demand_60_days": (days_from_as_of > 30) & (days_from_as_of <= 60),
            "demand_90_days": (days_from_as_of > 60) & (days_from_as_of <= 90),
            "demand_180_days": (days_from_as_of > 90) & (days_from_as_of <= 180),
            "demand_360_days": (days_from_as_of > 180) & (days_from_as_of <= 360),
            "demand_greater_than_360_day": days_from_as_of > 360,
        }

    for bucket_name, mask in bucket_masks.items():
        bucket = df[mask].groupby("item_number", dropna=False)["demand_qty"].sum().rename(bucket_name)
        summary = summary.join(bucket, how="outer")

    return summary.fillna(0).reset_index()[columns]

def build_location_lookup(location_df: pd.DataFrame) -> dict[str, list[tuple[str, float]]]:
    lookup: dict[str, list[tuple[str, float]]] = {}
    if location_df.empty:
        return lookup

    for item_number, rows in location_df.groupby("item_number", dropna=False):
        lookup[str(item_number)] = [
            (str(row.location_id), float(row.location_qty))
            for row in rows.sort_values("location_id", kind="stable").itertuples(index=False)
        ]
    return lookup


def format_allocated_qty(qty: float) -> str:
    return f"{qty:.0f}"


def allocate_location_string(location_lookup: dict[str, list[tuple[str, float]]], item_number: str, bucket_qty: float) -> str:
    remaining = float(bucket_qty or 0)
    if remaining <= 0:
        return ""

    parts: list[str] = []
    for location_id, location_qty in location_lookup.get(str(item_number), []):
        if remaining <= 0:
            break
        allocated_qty = min(location_qty, remaining)
        if allocated_qty > 0:
            parts.append(f"{location_id} * {format_allocated_qty(allocated_qty)}")
            remaining -= allocated_qty
    return "; ".join(parts)


def build_balance(stored_item_df: pd.DataFrame, demand_df: pd.DataFrame, as_of_date: date, bucket_mode: str) -> pd.DataFrame:
    inventory_summary, location_summary = build_inventory_summary(stored_item_df)
    demand_summary = build_demand_summary(demand_df, as_of_date, bucket_mode)
    location_lookup = build_location_lookup(location_summary)

    item_values = inventory_summary["item_number"].astype("string").dropna().astype(str).unique().tolist()
    balance = pd.DataFrame({"item": sorted(item_values)})
    balance = balance.merge(inventory_summary, left_on="item", right_on="item_number", how="left")
    balance = balance.merge(demand_summary, left_on="item", right_on="item_number", how="left", suffixes=("", "_demand"))
    balance = balance.drop(columns=[col for col in ["item_number", "item_number_demand"] if col in balance.columns])

    demand_cols = ["all_demand_qty"] + [f"demand_{days}_days" for days in BUCKET_DAYS] + ["demand_greater_than_360_day"]
    numeric_cols = ["sto_qty", "shipping_stage_qty"] + demand_cols
    for col in numeric_cols:
        if col not in balance.columns:
            balance[col] = 0
        balance[col] = clean_numeric(balance[col])

    # AS400 OPEN_CO_QTY is open order demand and is not reduced by HJ shipping stage quantity.
    balance["no_demand_qty"] = (balance["sto_qty"] - balance["all_demand_qty"]).clip(lower=0)
    for days in BUCKET_DAYS:
        demand_col = f"demand_{days}_days"
        no_demand_col = f"over_{days}_days_no_demand_qty"
        balance[no_demand_col] = (balance["sto_qty"] - balance[demand_col]).clip(lower=0)

    allocation_targets = [("no_demand_qty", "no_demand_allocated_location_qty")]
    allocation_targets.extend((f"over_{days}_days_no_demand_qty", f"over_{days}_days_allocated_location_qty") for days in BUCKET_DAYS)
    for qty_col, output_col in allocation_targets:
        balance[output_col] = [allocate_location_string(location_lookup, item, qty) for item, qty in zip(balance["item"], balance[qty_col])]

    ordered_columns = [
        "item",
        "sto_qty",
        "shipping_stage_qty",
        "all_demand_qty",
        "no_demand_qty",
        "no_demand_allocated_location_qty",
        "demand_30_days",
        "over_30_days_no_demand_qty",
        "over_30_days_allocated_location_qty",
        "demand_60_days",
        "over_60_days_no_demand_qty",
        "over_60_days_allocated_location_qty",
        "demand_90_days",
        "over_90_days_no_demand_qty",
        "over_90_days_allocated_location_qty",
        "demand_180_days",
        "over_180_days_no_demand_qty",
        "over_180_days_allocated_location_qty",
        "demand_360_days",
        "demand_greater_than_360_day",
        "over_360_days_no_demand_qty",
        "over_360_days_allocated_location_qty",
    ]
    nonzero_mask = (balance["sto_qty"] != 0) | (balance["shipping_stage_qty"] != 0) | (balance["all_demand_qty"] != 0)
    return balance.loc[nonzero_mask, ordered_columns].reset_index(drop=True)


def add_demand_bucket_columns(demand_df: pd.DataFrame, as_of_date: date, bucket_mode: str) -> pd.DataFrame:
    if demand_df.empty:
        return demand_df

    df = demand_df.copy()
    as_of_timestamp = pd.Timestamp(as_of_date)
    df["days_from_as_of"] = (df["demand_date"] - as_of_timestamp).dt.days
    if bucket_mode == "cumulative":
        df["demand_bucket"] = "360_9999_days"
        df.loc[df["demand_date"].isna(), "demand_bucket"] = "0_30_days_or_no_date"
        df.loc[df["days_from_as_of"] <= 30, "demand_bucket"] = "0_30_days"
        df.loc[(df["days_from_as_of"] > 30) & (df["days_from_as_of"] <= 60), "demand_bucket"] = "0_60_days"
        df.loc[(df["days_from_as_of"] > 60) & (df["days_from_as_of"] <= 90), "demand_bucket"] = "0_90_days"
        df.loc[(df["days_from_as_of"] > 90) & (df["days_from_as_of"] <= 180), "demand_bucket"] = "0_180_days"
        df.loc[(df["days_from_as_of"] > 180) & (df["days_from_as_of"] <= 360), "demand_bucket"] = "0_360_days"
    else:
        df["demand_bucket"] = "greater_than_360_days"
        df.loc[df["demand_date"].isna(), "demand_bucket"] = "0_30_days_or_no_date"
        df.loc[df["days_from_as_of"] <= 30, "demand_bucket"] = "0_30_days"
        df.loc[(df["days_from_as_of"] > 30) & (df["days_from_as_of"] <= 60), "demand_bucket"] = "31_60_days"
        df.loc[(df["days_from_as_of"] > 60) & (df["days_from_as_of"] <= 90), "demand_bucket"] = "61_90_days"
        df.loc[(df["days_from_as_of"] > 90) & (df["days_from_as_of"] <= 180), "demand_bucket"] = "91_180_days"
        df.loc[(df["days_from_as_of"] > 180) & (df["days_from_as_of"] <= 360), "demand_bucket"] = "181_360_days"
    return df

def autofit_worksheet_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        header = str(column_cells[0].value)
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length, len(header)) + 2, 60)


def apply_balance_bucket_colors(worksheet) -> None:
    bucket_fills = {
        "all": PatternFill("solid", fgColor="D9EAF7"),
        "30": PatternFill("solid", fgColor="DDEED6"),
        "60": PatternFill("solid", fgColor="FFF2CC"),
        "90": PatternFill("solid", fgColor="FCE4D6"),
        "180": PatternFill("solid", fgColor="EADCF8"),
        "360": PatternFill("solid", fgColor="DDEBF7"),
        "gt360": PatternFill("solid", fgColor="F4CCCC"),
    }
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_to_bucket = {
        "all_demand_qty": "all",
        "no_demand_qty": "all",
        "no_demand_allocated_location_qty": "all",
        "demand_30_days": "30",
        "over_30_days_no_demand_qty": "30",
        "over_30_days_allocated_location_qty": "30",
        "demand_60_days": "60",
        "over_60_days_no_demand_qty": "60",
        "over_60_days_allocated_location_qty": "60",
        "demand_90_days": "90",
        "over_90_days_no_demand_qty": "90",
        "over_90_days_allocated_location_qty": "90",
        "demand_180_days": "180",
        "over_180_days_no_demand_qty": "180",
        "over_180_days_allocated_location_qty": "180",
        "demand_360_days": "360",
        "over_360_days_no_demand_qty": "360",
        "over_360_days_allocated_location_qty": "360",
        "demand_greater_than_360_day": "gt360",
    }

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for column_cells in worksheet.iter_cols(min_row=1, max_row=worksheet.max_row):
        header = column_cells[0].value
        fill = bucket_fills.get(header_to_bucket.get(header))
        if fill is None:
            continue
        for cell in column_cells[1:]:
            cell.fill = fill

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def as400_demand_export_view(demand_df: pd.DataFrame) -> pd.DataFrame:
    requested_columns = [
        "ORDNO",
        "ITMSQ",
        "CCUSNO",
        "CUSNM",
        "CSHPNO",
        "BDTRP#",
        "TRIP_QTY",
        "HOUSE",
        "ITNBR",
        "OPEN_CO_QTY",
        "LOADDATE",
    ]
    return demand_df[[col for col in requested_columns if col in demand_df.columns]].copy()


def export_result(balance_df: pd.DataFrame, stored_item_df: pd.DataFrame, demand_df: pd.DataFrame, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    demand_export_df = as400_demand_export_view(demand_df)

    if output_path.suffix.lower() == ".csv":
        balance_df.to_csv(output_path, index=False, encoding="utf-8-sig")
        stored_item_df.to_csv(output_path.with_name(f"{output_path.stem}_stored_item_detail.csv"), index=False, encoding="utf-8-sig")
        demand_export_df.to_csv(output_path.with_name(f"{output_path.stem}_as400_demand_detail.csv"), index=False, encoding="utf-8-sig")
        return

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        balance_df.to_excel(writer, sheet_name="UPH_Balance", index=False)
        stored_item_df.to_excel(writer, sheet_name="Stored_Item_Detail", index=False)
        demand_export_df.to_excel(writer, sheet_name="AS400_Demand_Detail", index=False)
        autofit_worksheet_columns(writer.sheets["UPH_Balance"])
        apply_balance_bucket_colors(writer.sheets["UPH_Balance"])
        autofit_worksheet_columns(writer.sheets["Stored_Item_Detail"])
        autofit_worksheet_columns(writer.sheets["AS400_Demand_Detail"])


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export UPH STO balance using HJ inventory and AS400 open-order demand.")
    parser.add_argument("--server", default=DEFAULT_SERVER, help="HighJump SQL Server name.")
    parser.add_argument("--database", default=DEFAULT_DATABASE, help="HighJump SQL Server database name.")
    parser.add_argument("--driver", default=DEFAULT_DRIVER, help="HighJump SQL Server ODBC driver name.")
    parser.add_argument("--wh-id", default=DEFAULT_WH_ID, help="HighJump warehouse id.")
    parser.add_argument("--pick-put-id", default=DEFAULT_PICK_PUT_ID, help="HighJump t_item_master.pick_put_id filter.")
    parser.add_argument("--encrypt", default=DEFAULT_ENCRYPT, choices=["yes", "no", "optional", "mandatory"], help="HighJump ODBC encryption setting.")
    parser.add_argument("--db2-dsn", default=DEFAULT_DB2_DSN, help="AS400 DB2 ODBC DSN.")
    parser.add_argument("--db2-uid", default=DEFAULT_DB2_UID, help="AS400 DB2 user id.")
    parser.add_argument("--db2-password", default=None, help="AS400 DB2 password. Defaults to DB2_PWD env var or prompt.")
    parser.add_argument("--db2-product", default="ALL", help="Deprecated; AS400 demand is no longer filtered by product.")
    parser.add_argument("--full-db2-demand-detail", action="store_true", help="Use db2_ashton_demand.sql full 44-column detail instead of the faster light query.")
    parser.add_argument("--bucket-mode", default="cumulative", choices=["cumulative", "exclusive"], help="Demand bucket logic: cumulative uses 0-30/0-60/0-90/0-180/0-360/360-9999; exclusive uses non-overlapping ranges.")
    parser.add_argument("--as-of-date", default=date.today().isoformat(), help="Demand bucket as-of date in YYYY-MM-DD.")
    parser.add_argument("--output", type=Path, default=default_output_path(), help="Output .xlsx or .csv path.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    as_of_date = datetime.strptime(args.as_of_date, "%Y-%m-%d").date()

    print(f"Connecting to HJ {args.server}/{args.database}, wh_id={args.wh_id}, pick_put_id={args.pick_put_id}...")
    stored_item_df = fetch_hj_stored_item_detail(args)

    detail_mode = "full db2_ashton_demand.sql" if args.full_db2_demand_detail else "light AS400 demand"
    print(f"Connecting to AS400 DB2 DSN={args.db2_dsn}, mode={detail_mode}, product filter=NONE...")
    demand_df = fetch_as400_demand_detail(args)
    demand_df = add_demand_bucket_columns(demand_df, as_of_date, args.bucket_mode)

    balance_df = build_balance(stored_item_df, demand_df, as_of_date, args.bucket_mode)
    export_result(balance_df, stored_item_df, demand_df, args.output)
    print(
        f"Exported {len(balance_df):,} balance rows, "
        f"{len(stored_item_df):,} stored-item detail rows, and "
        f"{len(demand_df):,} AS400 demand detail rows to {args.output}"
    )


if __name__ == "__main__":
    main()



















