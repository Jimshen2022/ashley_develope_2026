import openpyxl
import pandas as pd
from pathlib import Path

# Read Excel file
excel_file = Path(r"D:\GitHub\Ashley_Project\Network Labor Planning File_DC Version.xlsx")
if excel_file.exists():
    print(f"✓ Found Excel file: {excel_file}")
    print(f"\nSheet names:")
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    for i, sheet in enumerate(wb.sheetnames[:20]):
        print(f"  {i}: {sheet}")
        
    # Try to find ADV HC Review sheet
    try:
        if 'ADV HC Review' in wb.sheetnames:
            print("\n✓ Found 'ADV HC Review' sheet")
            ws = wb['ADV HC Review']
            
            # Get first 30 rows to understand structure
            print("\nFirst 30 rows (columns A-J):")
            for row in range(1, 31):
                row_data = []
                for col in range(1, 11):  # A to J
                    cell = ws.cell(row, col)
                    row_data.append(f"{cell.value}")
                print(f"Row {row}: {row_data}")
        else:
            print("\n⚠ 'ADV HC Review' sheet not found")
            
    except Exception as e:
        print(f"Error accessing sheet: {e}")
else:
    print(f"✗ File not found: {excel_file}")
