from openpyxl import Workbook
from openpyxl import load_workbook
import os

root = os.getcwd()
wb = load_workbook(r'd:\python_file\AshtonOH.xlsx')
wb.create_sheet('Summary')
sht = wb['Summary']

sht.cell(row=1, column=1).value = '部门'

# 复制表头到summary Sheet
sht1 = wb.worksheets[1]
min_col = sht1.min_column
max_col = sht1.max_column
for i in range(min_col, max_col):
    sht.cell(row=1, column=i + 1).value = sht1.cell(row=1, column=i).value

# 遍历除"summary"工作表以外的每个工作表
for sht2 in wb.worksheets:
    if sht2.title != 'Summary':
        # "Summary" 工作表数据区域下面第一个空行, 加在这里循环是因为复制完一个表后，要将summary的最大行更新
        max_row0 = sht.max_row + 1
        # 要复制的sht2工作表的数据范围
        min_col = sht2.min_column
        max_col = sht2.max_column
        min_row = sht2.min_row + 1
        max_row = sht2.max_row

        # 复制数据
        n = 0
        for i in range(min_row, max_row + 1):
            n += 1
            for j in range(min_col, max_col + 1):
                sht.cell(row=max_row0 + n - 1, column=j + 1).value = \
                    sht2.cell(row=i, column=j).value
        # 在第1列添加部门名称
        rows0 = max_row - min_row + 1  # 要合并的资料区行数
        max_row1 = max_row0 + rows0 - 1
        # 以上max_row1为summary行+要合并资料区行数后的最大行。 （-1是因为max_ro0 = sht.max_ro+1）
        for i in range(max_row0, max_row1 + 1):
            sht.cell(row=i, column=1).value = sht2.title
wb.save(r'd:\python_file\AshtonOH9.xlsx')










