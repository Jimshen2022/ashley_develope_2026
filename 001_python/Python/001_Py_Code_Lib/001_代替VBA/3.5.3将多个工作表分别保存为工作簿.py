from openpyxl import load_workbook
from openpyxl import Workbook
import os

root = os.getcwd()
wb = load_workbook(r'd:\python_file\AshtonOH.xlsx')
for sht in wb.worksheets:
    row_min = sht.min_row
    row_max = sht.max_row
    col_min = sht.min_column
    col_max = sht.max_column
    wb1 = Workbook()
    sht1 = wb1.active
    for i in range(row_min,row_max+1):
        for j in range(col_min,col_max):
            sht1.cell(row=i,column=j).value=\
                sht.cell(row=i,column=j).value
    wb1.save('d:\\python_file\\'+sht.title+'.xlsx')
    wb1.close()
