from openpyxl import load_workbook
import os

root = r'd:\python_file'
wb = load_workbook(root + '\\AshtonOH3.xlsx')
shts = wb.worksheets
# 获取sheet1工作表
sht = wb.worksheets[0]
# sht = wb['Sheet1']
irow = sht.max_row  # 获取数据行数
strs = []
for i in range(2, irow + 1):  # 遍历每行数据
    strt = sht.cell(row=i, column=3).value  # 获取该行所属部门名称
    if (strt not in strs):
        # 如果是新部门, 则将名称添加到strs列表中
        strs.append(strt)
        sht1 = wb.create_sheet(strt)  # 新建工作表
        for j in range(1, sht.max_column):  # 为新工作表添加表头
            sht1.cell(row=1, column=j).value = \
                sht.cell(row=1, column=j).value
            sht1.cell(row=2, column=j).value = \
                sht.cell(row=i, column=j).value
    else:
        # 如果是已存在的部门名称, 则直接追加数据行
        r = wb[strt].max_row + 1
        for j in range(1, sht.max_column):  # 追加数据行
            wb[strt].cell(row=r, column=j).value = \
                sht.cell(row=i, column=j).value

# 删除新生成的工作表的第3列
shts = wb.worksheets
for k in range(1, len(shts)):
    sht = shts[k]
    if (sht.title != 'Sheet1'):
        sht.delete_cols(3)

# 删除 除Sheet1 的工作表
print(len(shts))
for m in range(len(shts)-1,0,-1):
    if (shts[m].title != 'Sheet1'):
        print(shts[m].title)
        # wb.remove(shts[m])
        del wb[shts[m].title]

wb.save(root + '\\AshtonOH7.xlsx')
