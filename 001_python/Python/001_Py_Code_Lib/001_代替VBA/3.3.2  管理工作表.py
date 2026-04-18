import openpyxl as pyxl

wb = pyxl.Workbook()
wb.create_sheet()
ws1 = wb.create_sheet('MySheet', 0)
sheets = wb.worksheets  # worksheet的集合
print(sheets[0].title)
print(len(sheets))

ws1 = wb.remove(ws1)    # 删除了 MySheet
sheets = wb.worksheets
print(sheets[0].title)
print(len(sheets))

for sheet in sheets:
    print(sheet.title)

