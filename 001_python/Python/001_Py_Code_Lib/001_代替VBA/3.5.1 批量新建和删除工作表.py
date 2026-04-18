# 批量新建工作表
# from openpyxl import Workbook
# import os
# root = os.getcwd()    # get current file path
# wb = Workbook()
# sht = wb.active
# for i in range(1,11):
#     wb.create_sheet(title='Sheet{}'.format(i),index=-1)
# wb.save(r'd:\python_file\CreationWb.xlsx')


# Remove方法批量删除工作表
from openpyxl import load_workbook
import os
root = r'd:\python_file'
wb = load_workbook(root + '\\CreationWb.xlsx')
for i in range(10,0,-1):
    wb.remove(wb.worksheets[i])
wb.save(root+'\\test.xlsx')







