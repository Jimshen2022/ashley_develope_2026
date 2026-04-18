from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\jishen\Downloads\ASSTO.xlsx')
ws0 = wb.create_sheet()  # 新增工作表名为sheet的工作表
ws1 = wb.create_sheet('Mysheet')  # 新增Mysheet工作表
ws2 = wb.create_sheet('Mysheet', 0)  # 0表示将此新增的sheet放到最前面
ws3 = wb.create_sheet('Mysheet', -1)  # -1表示将此新增的sheet放到倒数第2位置
sheets = wb.worksheets
for i in sheets:
    print(i.title)  # title打印sheet的名称

wb.remove(ws2)  # 删除
del wb[ws3.title]
print('-------------------------------------------------------------------------------------')
sheets2 = wb.worksheets
for i in sheets2:
    print(i.title)  # title打印sheet的名称

ws4 = sheets2[0]
print(ws4.title)  # 工作表的名称

print('-------------------------------------------------------------------------------------')
print(ws4.sheet_state)  # 可见状态
print(ws4.dimensions)  # 表格中含有数据部分的大小
print(ws4.sheet_properties)  # 表格中含有数据部分的大小
ws4.sheet_properties.tabColor = 'FF0000'  # 设置选项卡标签处的背景色
print(ws4.active_cell)    # 活动单元格
print(ws4.selected_cell)    # 选中的单元格
