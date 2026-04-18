
# 创建和删除工作表
import openpyxl as pyxl
wb = pyxl.Workbook()                # 会自动产生一个工作表为Sheet的sheet1
ws0 = wb.create_sheet()            # 会自动产生一个工作表为Sheet1的Sheet2
ws1 = wb.create_sheet('MySheet')    # 会自动产生一个工作表为 MySheet 的Sheet3
ws2 = wb.create_sheet('MySheet',0)    # 会自动产生一个工作表为 MySheet1 的Sheet1--会在最前面, 其余工作表名按此+1
ws3 = wb.create_sheet('MySheet',-1)   # 倒数第2插入一张工作表为 Mysheet2， 并且所有表从后往前编号， 负数表示从右往左编号工作表
# ws4 = wb.create_sheet('MySheet',0)
ws4 = wb.active           # 获取活动工作表
print(ws4.title)          # 活动工作表名
wb.save(r'd:\python_file\openpyxl.xlsx')


wb2 = pyxl.Workbook()
ws5 = wb2.active
print(ws5.title)

wb.remove(ws1)
del wb[ws1.title]




















