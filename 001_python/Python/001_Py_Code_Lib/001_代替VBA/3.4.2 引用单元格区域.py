from openpyxl import load_workbook

wb = load_workbook(r'd:\python_file\AshtonOH.xlsx')
ws = wb.worksheets
for sht in ws:
    print(sht.title)
ws1 = ws[0]
cr = ws1['a1:c4']  # 得到元组
print(cr[2][2].value)  # 取元组某个单元格的值

# cellRange对象的主要属性和方法:
from openpyxl.worksheet import cell_range as cr
cr0 = cr.CellRange(min_row=2, max_row=5, min_col=3, max_col=6)

print(cr0.bottom)    # 区域底部一行各单元格的坐标
print(cr0.top)    # 区域顶部一行各单元格的坐标
print(cr0.left)    # 区域左侧一列各单元格的坐标
print(cr0.right)    # 区域右侧一列各单元格的坐标
print(cr0.min_row)    # 区域最小行号
print(cr0.min_col)    # 区域最小列号
print(cr0.max_row)    # 区域最大行号
print(cr0.max_col)    # 区域最大列号
print(cr0.size)    # 区域大小
print(cr0.bounds)    # 区域左上角和右下角单元格的坐标
print(cr0.coord)    # 区域左上角和右下角单元格的坐标
for cell in cr0.rows:    # 区域各行单元格的坐标
    print(cell)
for cell2 in cr0.cols:    # 区域各列单元格的坐标
    print(cell2)
for cell3 in cr0.cells:    # 区域各单元格的坐标
    print(cell3)





