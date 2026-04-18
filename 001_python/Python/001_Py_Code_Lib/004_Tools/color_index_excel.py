import matplotlib.pyplot as plt
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Matplotlib 默认颜色
default_colors = plt.rcParams['axes.prop_cycle'].by_key()['color']

# SCI 常用颜色（手工精选，参考 Nature / Cell / Science 等）
sci_colors = [
    "#E41A1C",  # red
    "#377EB8",  # blue
    "#4DAF4A",  # green
    "#984EA3",  # purple
    "#FF7F00",  # orange
    "#A65628",  # brown
    "#F781BF",  # pink
    "#999999",  # grey
    "#66C2A5",  # teal
    "#FC8D62",  # coral
    "#8DA0CB",  # lavender
    "#E78AC3",  # pink
    "#A6D854",  # light green
    "#FFD92F",  # yellow
]

# 合并颜色列表（Matplotlib + SCI）
all_colors = default_colors + sci_colors

# 生成时间戳
timestamp = datetime.now().strftime('%Y%m%d_%H%M')
filename = f'color_index_{timestamp}.xlsx'

# 下载路径
downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
filepath = os.path.join(downloads_folder, filename)

# 创建 Excel 文件
wb = Workbook()
ws = wb.active
ws.title = "Color Index"

# 写入标题
ws.append(["Index", "Color (Hex)", "Preview", "Source"])

# 写入颜色数据
for idx, color in enumerate(all_colors):
    # 判断来源
    source = "Matplotlib default" if idx < len(default_colors) else "SCI common"

    # 写入值
    ws.cell(row=idx + 2, column=1, value=idx)         # A列: Index
    ws.cell(row=idx + 2, column=2, value=color)       # B列: 颜色 HEX
    ws.cell(row=idx + 2, column=4, value=source)      # D列: 来源

    # C列: 背景填色
    fill_color = color.replace("#", "")  # openpyxl 需要无井号的颜色代码
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    cell_preview = ws.cell(row=idx + 2, column=3)
    cell_preview.fill = fill

# 保存文件
wb.save(filepath)

print(f"颜色索引 Excel 文件已保存至: {filepath}")
