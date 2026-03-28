import time
import os
import shutil
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# Import all modules
import a01_pull_open_trips
import a021_pull_firmed_planned_po
import a03_pull_hj_sto
import a04_pull_out_yard
import a05_pull_out_open_po
import a06_yard_add_weeksaturday
import a07_add_condition_on_open_po
import a08_item_master
import a09_products
import a10_item_unique_list
import a11_inbound_cal
import a12_outbound_cal
import a13_cubes_unique_item
import a14_cubes_added_for_sheets
import a15_cubes_added_for_in_out
import summary  # a16

def clear_and_set_filters(wb):
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.auto_filter.ref = None
        if ws.max_row > 0 and ws.max_column > 0:
            ws.auto_filter.ref = ws.dimensions

def apply_summary_to_excel(wb, dfs):
    print("Applying calculated summary to 'Capacity vs Demand vs Supply'...")
    if 'Summary_Calculations' not in dfs: return
    calc = dfs['Summary_Calculations']
    
    ws_cap = wb['Capacity vs Demand vs Supply']
    dates = calc['dates']
    
    # Write dynamic dates
    for c in range(8):
        ws_cap.cell(row=2, column=5+c).value = dates[c]
        
    dict_rows = {}
    
    for r in range(3, ws_cap.max_row + 1):
        # 确保读取的是 B 列（Sub_Category）进行匹配，与 VBA 一致
        cat_cell = ws_cap.cell(row=r, column=2).value
        type_cell = ws_cap.cell(row=r, column=4).value
        
        curCat = str(cat_cell).strip().upper() if cat_cell else ""
        typeStr = str(type_cell).strip() if type_cell else ""
        
        if typeStr:
            cleanType = typeStr.replace(" ", "").replace("-", "").replace("_", "").upper()
            baseKey = ""
            
            if "CAPACITYBALANCE" in cleanType: baseKey = "CAPACITYBALANCE"
            elif "CAPACITY" in cleanType: baseKey = "CAPACITY"
            elif "ONHAND" in cleanType: baseKey = "ONHAND"
            elif "YARD" in cleanType: baseKey = "YARD"
            elif "INTRANSIT" in cleanType: baseKey = "INTRANSIT"
            elif "OPENPO" in cleanType: baseKey = "OPENPO"
            elif "FIRMED" in cleanType: baseKey = "FIRMED"
            elif "PLANNED" in cleanType: baseKey = "PLANNED"
            elif "TRIPPED" in cleanType: baseKey = "TRIPPED"
            elif "OPENCO" in cleanType: baseKey = "OPENCO"
            elif "DSBALANCE" in cleanType: baseKey = "DSBALANCE"
            elif "UTILIZATION" in cleanType: baseKey = "UTILIZATION"

            if curCat != "ALL" and baseKey != "":
                if baseKey not in ["CAPACITYBALANCE", "DSBALANCE", "UTILIZATION"]:
                    if baseKey not in dict_rows: dict_rows[baseKey] = []
                    dict_rows[baseKey].append(r)

            for c in range(8):
                dateVal = dates[c]
                dt_str = pd.to_datetime(dateVal).strftime('%Y-%m-%d')
                
                # 完美的字典键值拼接 (例如 "BULK|2026-03-28")
                dKey = f"{curCat}|{dt_str}"
                
                cell = ws_cap.cell(row=r, column=5+c)
                cell.number_format = '#,##0'
                col_letter = chr(69 + c) # 69 is 'E'
                
                if baseKey == "CAPACITYBALANCE":
                    cell.value = f"={col_letter}{r-10}-{col_letter}{r-1}"
                elif baseKey == "DSBALANCE":
                    if c == 0:
                        cell.value = f"={col_letter}{r-8}+{col_letter}{r-7}+{col_letter}{r-6}+{col_letter}{r-5}+{col_letter}{r-3}-{col_letter}{r-2}-{col_letter}{r-1}"
                    else:
                        prev_col_letter = chr(69 + c - 1)
                        cell.value = f"={prev_col_letter}{r}+{col_letter}{r-7}+{col_letter}{r-6}+{col_letter}{r-5}+{col_letter}{r-3}-{col_letter}{r-2}-{col_letter}{r-1}"
                elif baseKey == "UTILIZATION":
                    cell.value = f"=IF({col_letter}{r-11}=0,0,{col_letter}{r-2}/{col_letter}{r-11})"
                    cell.number_format = '0%'
                elif curCat == "ALL":
                    if baseKey == "ONHAND" and c > 0:
                        pass # Skip
                    else:
                        if baseKey in dict_rows and len(dict_rows[baseKey]) > 0:
                            formula_str = "=" + "+".join([f"{col_letter}{row_ref}" for row_ref in dict_rows[baseKey]])
                            cell.value = formula_str
                else:
                    if baseKey == "CAPACITY": cell.value = round(calc['dictCap'].get(curCat, 0), 0)
                    elif baseKey == "ONHAND": 
                        if c == 0: cell.value = calc['dictOH'].get(curCat, 0)
                    elif baseKey == "YARD": cell.value = calc['dictYard'].get(dKey, 0)
                    elif baseKey == "INTRANSIT": cell.value = calc['dictInTransit'].get(dKey, 0)
                    elif baseKey == "OPENPO": cell.value = calc['dictPO'].get(dKey, 0)
                    elif baseKey == "FIRMED": cell.value = calc['dictFirmed'].get(dKey, 0)
                    elif baseKey == "PLANNED": cell.value = calc['dictPlanned'].get(dKey, 0)
                    elif baseKey == "TRIPPED": cell.value = calc['dictTripped'].get(dKey, 0)
                    elif baseKey == "OPENCO": cell.value = calc['dictOpenCO'].get(dKey, 0)

def write_df_to_sheet(wb, df, sheet_name, start_row=1, start_col=1, clear_all=True, header=True):
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    ws = wb[sheet_name]
    
    if clear_all:
        ws.delete_rows(start_row, ws.max_row)
        
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=header), start_row):
        for c_idx, value in enumerate(row, start_col):
            # Format text properly to prevent scientific notation dropping 0s
            if c_idx == 1 and sheet_name in ['ItemBalance_by_Pieces', 'ItemBalance_by_Cubes']:
                cell = ws.cell(row=r_idx, column=c_idx, value=str(value))
                cell.number_format = '@'
            else:
                ws.cell(row=r_idx, column=c_idx, value=value)

def save_as_xlsb_and_upload(xlsx_path, target_dir):
    try:
        import win32com.client
    except ImportError:
        print("未安装 pywin32，跳过 xlsb 转换。")
        return None

    print(f"\n正在将文件格式转换为 .xlsb 格式...")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    
    # 修改生成的文件名规则
    xlsb_filename = f"Ashton Inbound & Outbound Balance - {timestamp}.xlsb"
    
    local_xlsb_path = os.path.join(os.path.dirname(xlsx_path), xlsb_filename)
    
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    
    try:
        wb = excel.Workbooks.Open(os.path.abspath(xlsx_path))
        wb.SaveAs(os.path.abspath(local_xlsb_path), FileFormat=50)
        wb.Close()
    except Exception as e:
        print(f"转换为 xlsb 失败: {e}")
        return None
    finally:
        excel.Quit()
        
    if target_dir:
        os.makedirs(target_dir, exist_ok=True)
        dest_path = os.path.join(target_dir, xlsb_filename)
        try:
            shutil.move(local_xlsb_path, dest_path)
            print(f"✅ 成功生成并移动到目标目录: {dest_path}")
            return dest_path
        except Exception as e:
            print(f"❌ 移动到目标目录失败: {e}")
            return local_xlsb_path
    else:
        print(f"✅ 成功生成文件: {local_xlsb_path}")
        return local_xlsb_path

def main(file_path, target_dir=""):
    print("开始执行自动化报表生成 (纯内存加速版)...")
    start_time = time.time()
    current_time_str = datetime.now().strftime("%Y/%m/%d %H:%M:%S")

    # The centralized memory dictionary
    dfs = {}

    # STEP 0: Optional pre-load some base sheets if needed (LocationCapacity, Setup)
    print("Pre-loading base configuration sheets...")
    try:
        xls = pd.ExcelFile(file_path)
        dfs['Setup'] = pd.read_excel(xls, sheet_name='Setup')
        if 'LocationCapacity' in xls.sheet_names:
            dfs['LocationCapacity'] = pd.read_excel(xls, sheet_name='LocationCapacity')
        if 'Capacity vs Demand vs Supply' in xls.sheet_names:
            dfs['Capacity vs Demand vs Supply'] = pd.read_excel(xls, sheet_name='Capacity vs Demand vs Supply')
    except Exception as e:
        print(f"Error reading base sheets: {e}")
        return

    # --- 数据处理模块调用区 (传递 dfs 字典进行内存操作) ---
    a01_pull_open_trips.run(dfs, file_path)
    a021_pull_firmed_planned_po.run(dfs)
    a03_pull_hj_sto.run(dfs)
    a04_pull_out_yard.run(dfs)
    a05_pull_out_open_po.run(dfs, file_path)
    a06_yard_add_weeksaturday.run(dfs)
    a07_add_condition_on_open_po.run(dfs)
    a08_item_master.run(dfs)
    a09_products.run(dfs)
    
    # 复杂聚合计算
    a10_item_unique_list.run(dfs)
    a11_inbound_cal.run(dfs)
    a12_outbound_cal.run(dfs)
    a13_cubes_unique_item.run(dfs)
    a14_cubes_added_for_sheets.run(dfs)
    a15_cubes_added_for_in_out.run(dfs)
    
    # Summary Dashboard 计算
    summary.update_capacity_demand_supply(dfs)
    
    print("内存计算完成，正在将所有结果一次性写入 Excel...")
    latest_file_path = file_path.replace('.xlsm', '_updated.xlsx')
    
    # Create a copy of the original file to safely write into
    shutil.copy2(file_path, latest_file_path)
    wb = openpyxl.load_workbook(latest_file_path)
    
    # Write memory dataframes back to their respective sheets
    sheet_mappings_df = {
        'CustomerOrder': dfs.get('CustomerOrder'),
        'Firmed_Planned_PO': dfs.get('Firmed_Planned_PO'),
        'OnHand': dfs.get('OnHand'),
        'Yard': dfs.get('Yard'),
        'OpenPO': dfs.get('OpenPO'),
        'Item_Master': dfs.get('Item_Master')
    }
    
    for sh_name, df in sheet_mappings_df.items():
        if df is not None and not df.empty:
            print(f"Writing {sh_name}...")
            write_df_to_sheet(wb, df, sh_name, start_row=1, start_col=1, clear_all=True, header=True)

    # Write Balance by Pieces
    if 'ItemBalance_by_Pieces_Data' in dfs:
        print("Writing ItemBalance_by_Pieces...")
        write_df_to_sheet(wb, dfs['ItemBalance_by_Pieces_Data'], 'ItemBalance_by_Pieces', start_row=6, start_col=1, clear_all=False, header=False)
        
    # Write Balance by Cubes
    if 'ItemBalance_by_Cubes_Data' in dfs:
        print("Writing ItemBalance_by_Cubes...")
        write_df_to_sheet(wb, dfs['ItemBalance_by_Cubes_Data'], 'ItemBalance_by_Cubes', start_row=6, start_col=1, clear_all=False, header=False)

    # Apply summary dashboard logic
    apply_summary_to_excel(wb, dfs)

    # Update Timestamps
    sheet_mapping = {
        'Sheet3': 'ItemBalance_by_Pieces',
        'Sheet4': 'ItemBalance_by_Cubes',
        'Sheet6': 'Capacity vs Demand vs Supply',
        'Sheet11': 'Setup'
    }
    
    try:
        from openpyxl.styles import Font
        red_font = Font(color="FF0000")
        
        if sheet_mapping['Sheet6'] in wb.sheetnames:
            ws6 = wb[sheet_mapping['Sheet6']]
            ws6['A1'].value = f"DataCollectedAt: {current_time_str}"
            ws6['A1'].font = red_font
        
        if sheet_mapping['Sheet3'] in wb.sheetnames:
            ws3 = wb[sheet_mapping['Sheet3']]
            ws3['A1'].value = f"DataCollectedAt: {current_time_str}"
            ws3['A1'].font = red_font
            ws3['B3'].value = "By Pieces:"
            # Set Target Dates in Row 4 for Inbound/Outbound
            if 'Target_Dates' in dfs:
                dates = dfs['Target_Dates']
                in_cols = [5, 8, 11, 14, 17, 20, 23]
                out_cols = [6, 9, 12, 15, 18, 21, 24]
                for i, d in enumerate(dates):
                    ws3.cell(row=4, column=in_cols[i]).value = d
                    ws3.cell(row=4, column=out_cols[i]).value = d

        if sheet_mapping['Sheet4'] in wb.sheetnames:
            ws4 = wb[sheet_mapping['Sheet4']]
            ws4['A1'].value = f"DataCollectedAt: {current_time_str}"
            ws4['A1'].font = red_font
            ws4['B3'].value = "By Cubes:"
            if 'Target_Dates' in dfs:
                dates = dfs['Target_Dates']
                in_cols = [5, 8, 11, 14, 17, 20, 23]
                out_cols = [6, 9, 12, 15, 18, 21, 24]
                for i, d in enumerate(dates):
                    ws4.cell(row=4, column=in_cols[i]).value = d
                    ws4.cell(row=4, column=out_cols[i]).value = d
            
        if sheet_mapping['Sheet11'] in wb.sheetnames:
            ws11 = wb[sheet_mapping['Sheet11']]
            max_row_g = 1
            for row in range(ws11.max_row, 0, -1):
                if ws11.cell(row=row, column=7).value is not None:
                    max_row_g = row
                    break
            
            elapsed_time = time.time() - start_time
            time_msg = f"Updated Successful~  Wall Time: {elapsed_time:,.2f}s."
            ws11.cell(row=max_row_g + 1, column=7).value = time_msg
        
    except Exception as e:
        print(f"更新时间戳时发生错误: {e}")

    clear_and_set_filters(wb)
    print("Saving intermediate .xlsx...")
    wb.save(latest_file_path)
    
    save_as_xlsb_and_upload(latest_file_path, target_dir)
    
    if os.path.exists(latest_file_path): 
        os.remove(latest_file_path)

    import gc
    dfs.clear()
    del dfs
    gc.collect()

    total_time = time.time() - start_time
    print(f"\n🎉 整个流程执行完毕! 总耗时: {total_time:,.2f} 秒.")

if __name__ == "__main__":
    source_file = r"/036_inbound_outbound_balance/054_inbound_outbound_capacity/Ashton Inbound & Outbound Balance - 20260324 1207pm.xlsm"
    downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
    
    main(source_file, downloads_folder)